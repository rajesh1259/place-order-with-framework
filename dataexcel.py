import openpyxl

class TestOpen:
    @staticmethod
    def test_myfun():
        sheet = openpyxl.load_workbook("/Volumes/MAC - Data/excel/selex.xlsx")
        current_sheet = sheet.active
        Dict = {}

        for i in range(1, current_sheet.max_row + 1):

            for j in range(7, current_sheet.max_column + 1):
                Dict[current_sheet.cell(row=1, column=j).value] = current_sheet.cell(row=3, column=j).value

        return[Dict]





